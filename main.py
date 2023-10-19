import openpyxl
import requests
from flask import Flask, request, jsonify, json

app = Flask(__name__)
# Путь к файлу CSVParcer.xlsx
excel_file_path = 'csv/CSVParcer.xlsx'

# Загрузка файла Excel
workbook = openpyxl.load_workbook(excel_file_path)

# Выбор активного листа (в данном случае первый лист)
sheet = workbook.active

# Создание пустых списков для данных PriceRunner и InventoryRunner
price_runner_data = {}
inventory_runner_data = {}




# Парсинг данных из листа PriceRunner
for row in sheet.iter_rows(min_row=34, max_row=42, values_only=True):
    parameter_number = row[2]
    parameter_name = row[3]
    parameter_value = row[4]
    parameter_description = row[5]
    price_runner_data[parameter_number] = {
        "name": parameter_name,
        "value": parameter_value,
        "description": parameter_description
    }

# Парсинг данных из листа InventoryRunner
for row in sheet.iter_rows(min_row=44, max_row=55, values_only=True):
    parameter_number = row[2]
    parameter_name = row[3]
    parameter_value = row[4]
    parameter_description = row[5]
    inventory_runner_data[parameter_number] = {
        "name": parameter_name,
        "value": parameter_value,
        "description": parameter_description
    }
#Преобразование json файлы
price_runner_json_data = json.dumps(price_runner_data, indent=4, ensure_ascii=False)
with open('price_runner_data.json', 'w', encoding='utf-8') as json_file:
    json_file.write(price_runner_json_data)

# Сохранение данных InventoryRunner в .json файл
inventory_runner_json_data = json.dumps(inventory_runner_data, indent=4, ensure_ascii=False)
with open('inventory_runner_data.json', 'w', encoding='utf-8') as json_file:
    json_file.write(inventory_runner_json_data)

#Пример вывода данных
print("PriceRunner data:")
print(price_runner_data)

print("\nInventoryRunner data:")
print(inventory_runner_data)

#Проверяем в Postman
@app.route('/inventoryUpdateBatch', methods=['POST'])
def inventory_update_batch():
    try:
        # Получаем данные JSON из тела POST-запроса
        data = request.get_json()

        # Обработка данных и метода inventoryUpdateBatch()
        # В данном примере просто возвращаем количество загруженных позиций
        items = data.get('items', [])
        count = len(items)

        # Подготавливаем ответ в формате JSON
        response = {
            'jsonrpc': '2.0',
            'id': data.get('id', ''),
            'result': {
                'count': count
            }
        }

        # Отправляем ответ клиенту
        return jsonify(response)

    except Exception as e:
        # В случае ошибки возвращаем соответствующий ответ
        error_response = {
            'jsonrpc': '2.0',
            'id': data.get('id', ''),
            'error': {
                'code': 500,
                'message': 'Internal Server Error'
            }
        }
        return jsonify(error_response), 500
if __name__ == '__main__':
    app.run(debug=True)
# задание согласно документации
def inventoryUpdateBatch(items, auth_key):
    # URL для отправки POST-запроса
    url = f"https://api.heado.ru/management/{auth_key}"

    # Подготовка данных для POST-запроса
    payload = {
        "id": "1251506945251332",
        "method": "inventoryUpdateBatch",
        "params": {
            "items": items
        },
        "jsonrpc": "2.0"
    }

    try:
        # Отправка POST-запроса
        response = requests.post(url, json=payload)

        # Печать ответа сервера
        print("Response:")
        print(response.json())

        # Возвращаем ответ сервера
        return response.json()

    except Exception as e:
        # В случае ошибки возвращаем соответствующий ответ
        error_response = {
            'jsonrpc': '2.0',
            'id': payload.get('id', ''),
            'error': {
                'code': 500,
                'message': f'Internal Server Error: {str(e)}'
            }
        }
        return error_response


# Чтение данных из файла price_runner_data.json
with open('price_runner_data.json', 'r') as json_file:
    # Считывание данных напрямую в список
    items = json.load(json_file)

# Замените на ваш реальный Auth Key. Этот к
auth_key = 'hPkRxt14oQ_9d3mXfQYolyTnPZP6hRFzFZpcrqhPggZujkqy9oXvXQTLOIS5ywEdu3pWu'

# Вызов метода для отправки данных
inventoryUpdateBatch(items, auth_key)

# задание согласно документации
def price_update_batch(auth_key, items):
    url = f'https://gamma.heado.ru/management/{auth_key}'
    payload = {
        "id": "1251506945251332",
        "method": "price.updateBatch",
        "params": {
            "items": items
        },
        "jsonrpc": "2.0"
    }
    headers = {'Content-Type': 'application/json'}

    try:
        # Отправка POST-запроса
        response = requests.post(url, json=payload)

        # Печать ответа сервера
        print("Response:")
        print(response.json())

        # Возвращаем ответ сервера
        return response.json()

    except Exception as e:
        # В случае ошибки возвращаем соответствующий ответ
        error_response = {
            'jsonrpc': '2.0',
            'id': payload.get('id', ''),
            'error': {
                'code': 500,
                'message': f'Internal Server Error: {str(e)}'
            }
        }
        return error_response



auth_key = 'hPkRxt14oQ_9d3mXfQYolyTnPZP6hRFzFZpcrqhPggZujkqy9oXvXQTLOIS5ywEdu3pWu'
with open('inventory_runner_data.json', 'r') as json_file:
    # Считывание данных напрямую в списо
    items = json.load(json_file)
price_update_batch(auth_key, items)



